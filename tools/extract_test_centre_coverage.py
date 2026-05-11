from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract Drivest test centre coverage from a route status workbook."
    )
    parser.add_argument("workbook", help="Path to the XLSX workbook")
    parser.add_argument(
        "--output",
        default="site/data/test-centre-coverage.en-GB.json",
        help="Output JSON path relative to the repo root",
    )
    parser.add_argument(
        "--min-routes-exclusive",
        type=int,
        default=2,
        help="Exclude centres with route counts less than or equal to this value",
    )
    return parser.parse_args()


def workbook_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["all_centres"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        yield row


def extract_coverage(workbook_path: Path, min_routes_exclusive: int) -> dict:
    centres = []
    status_counter: Counter[str] = Counter()

    for row in workbook_rows(workbook_path):
        name = row[0]
        centre_id = row[1]
        status = row[2]
        route_count = row[3]

        if route_count is None:
            continue

        try:
            route_count_int = int(route_count)
        except (TypeError, ValueError):
            continue

        if route_count_int <= min_routes_exclusive:
            continue

        status_text = str(status or "").strip() or "unknown"
        status_counter[status_text] += 1
        centres.append(
            {
                "name": str(name).strip(),
                "id": str(centre_id).strip(),
                "routeCount": route_count_int,
                "status": status_text,
            }
        )

    centres.sort(key=lambda item: (item["name"].lower(), item["id"]))

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": workbook_path.name,
        "filter": {
            "routeCountGreaterThan": min_routes_exclusive,
        },
        "summary": {
            "centres": len(centres),
            "routes": sum(item["routeCount"] for item in centres),
            "statusCounts": dict(sorted(status_counter.items())),
        },
        "centres": centres,
    }


def main() -> None:
    args = parse_args()
    repo_root = Path(__file__).resolve().parent.parent
    workbook_path = Path(args.workbook).expanduser().resolve()
    output_path = (repo_root / args.output).resolve()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = extract_coverage(workbook_path, args.min_routes_exclusive)
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
